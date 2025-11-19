import csv
import datetime
from openpyxl import Workbook


def calculate_age(birth_date_str):
    # Розрахунок повних років
    birth_date = datetime.datetime.strptime(birth_date_str, '%Y-%m-%d').date()
    today = datetime.date.today()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age


def create_xlsx_from_csv(csv_filename, xlsx_filename):
    try:
        wb = Workbook()

        # Створення аркушів
        ws_all = wb.active
        ws_all.title = "all"
        ws_younger_18 = wb.create_sheet("younger_18")
        ws_18_45 = wb.create_sheet("18-45")
        ws_45_70 = wb.create_sheet("45-70")
        ws_older_70 = wb.create_sheet("older_70")

        sheets_map = {
            "all": ws_all,
            "younger_18": ws_younger_18,
            "18-45": ws_18_45,
            "45-70": ws_45_70,
            "older_70": ws_older_70
        }

        with open(csv_filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file, delimiter=';')
            headers = next(reader)

            full_headers = headers + ['Вік']

            for ws in sheets_map.values():
                ws.append(full_headers)

            for row in reader:
                birth_date_str = row[4]  # Індекс 4 - дата народження
                age = calculate_age(birth_date_str)

                row_with_age = row + [age]

                # Запис в загальний аркуш
                ws_all.append(row_with_age)

                # Розподіл за категоріями
                if age < 18:
                    ws_younger_18.append(row_with_age)
                elif 18 <= age <= 45:
                    ws_18_45.append(row_with_age)
                elif 45 < age <= 70:
                    ws_45_70.append(row_with_age)
                else:  # older than 70
                    ws_older_70.append(row_with_age)

        wb.save(xlsx_filename)
        print("Ok")

    except FileNotFoundError:
        print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
    except Exception as e:
        print(f"Повідомлення про неможливість створення XLSX файлу: {e}")


if __name__ == "__main__":
    create_xlsx_from_csv('employees.csv', 'employees.xlsx')